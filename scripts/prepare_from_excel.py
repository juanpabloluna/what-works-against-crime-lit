#!/usr/bin/env python3
"""
Prepare ingestion CSV from Excel citation database + PDF folder tree.

Reads the 'Base de datos review' Excel file, parses APA citations to extract
metadata, recursively scans the PDF folder, and matches citations to PDFs
by year + first author last name.

Usage:
    python scripts/prepare_from_excel.py \
        --excel "/Users/jpl/Downloads/what works against crime/Base de datos review (compartido).xlsx" \
        --pdfs "/Users/jpl/Downloads/what works against crime/data set of studies" \
        --output data/papers_metadata.csv
"""

import argparse
import csv
import re
import sys
from pathlib import Path
from collections import defaultdict


def parse_citation(raw: str) -> dict:
    """Parse an APA-style citation string into structured fields."""
    raw = raw.strip()

    # Extract DOI if present
    doi = ""
    doi_match = re.search(r"https?://doi\.org/(10\.\S+?)[\s\)]*$", raw)
    if not doi_match:
        doi_match = re.search(r"(10\.\d{4,}/\S+?)[\s\)]*$", raw)
    if doi_match:
        doi = doi_match.group(1).rstrip(".")

    # Extract year: look for (YYYY)
    year = ""
    year_match = re.search(r"\((\d{4})\)", raw)
    if year_match:
        year = year_match.group(1)

    # Extract authors: everything before the first (YYYY)
    authors_raw = ""
    if year_match:
        authors_raw = raw[: year_match.start()].strip().rstrip(",").strip()

    # Parse authors into semicolon-separated format
    authors = _parse_authors(authors_raw)

    # Extract title: text between (YYYY). and the next period that precedes a journal name
    title = ""
    if year_match:
        after_year = raw[year_match.end() :].strip()
        # Remove leading ". " or ".) " etc.
        after_year = re.sub(r"^[\.\)\s]+", "", after_year)
        # Title is typically the first sentence (ending with . or :)
        # But titles can contain periods, so look for the pattern "Title. Journal,"
        title_match = re.match(r"(.+?)\.\s+([A-Z])", after_year)
        if title_match:
            title = title_match.group(1).strip()
        else:
            # Fallback: take everything up to the first period
            parts = after_year.split(".", 1)
            title = parts[0].strip() if parts else after_year

    # Extract journal: text after title, before volume/issue
    publication = ""
    if title and year_match:
        after_year = raw[year_match.end() :].strip()
        after_year = re.sub(r"^[\.\)\s]+", "", after_year)
        after_title = after_year[len(title) :].strip().lstrip(".")
        # Journal is typically before the first comma followed by volume number
        journal_match = re.match(r"\s*(.+?),\s*\d", after_title)
        if journal_match:
            publication = journal_match.group(1).strip()

    return {
        "authors": authors,
        "year": year,
        "title": title,
        "publication": publication,
        "doi": doi,
    }


def _parse_authors(raw: str) -> str:
    """Convert APA author string to semicolon-separated 'Last, First' format."""
    if not raw:
        return ""
    # Normalize separators: ", &" and " & " become author boundaries
    # APA format: "LastA, F. A., LastB, F. B., & LastC, F. C."
    # Strategy: split on ", & " or " & " first, then split remaining on
    # the pattern ", Capital" which indicates a new author
    raw = raw.replace(", &", ",&").replace(" &", ",&")
    parts = [a.strip().lstrip("&").strip() for a in raw.split(",&") if a.strip()]

    # Each part might still be "LastA, F. A., LastB, F. B."
    # Split on the pattern where a comma-space is followed by an uppercase letter
    # that looks like a new last name (not an initial like "F.")
    final_authors = []
    for part in parts:
        # Split "LastA, F. A., LastB, F. B." into individual authors
        # Pattern: split before "Uppercase-word, " that follows a period or initial
        sub_authors = re.split(r",\s+(?=[A-Z][a-z]{2,})", part)
        for sa in sub_authors:
            sa = sa.strip().rstrip(",").strip()
            if sa:
                final_authors.append(sa)

    return "; ".join(final_authors)


def find_pdfs(pdf_root: Path) -> dict:
    """Recursively find all PDFs and index by (year, normalized_author)."""
    pdf_index = defaultdict(list)  # (year, author_key) -> [path, ...]
    all_pdfs = []

    for pdf_path in sorted(pdf_root.rglob("*.pdf")):
        fname = pdf_path.stem  # e.g. "2012 Braga & Weisburd"
        all_pdfs.append(pdf_path)

        # Parse year and first author from filename
        m = re.match(r"(\d{4})\w?\s+(.+)", fname)
        if m:
            yr = m.group(1)
            author_part = m.group(2)
            # First author: take text before &, et al, comma
            first_author = re.split(r"[&,]|\bet al\b", author_part)[0].strip()
            key = (yr, first_author.lower())
            pdf_index[key].append(pdf_path)

    return pdf_index, all_pdfs


def match_citation_to_pdf(citation: dict, pdf_index: dict, all_pdfs: list):
    """Try to match a parsed citation to a PDF file."""
    year = citation["year"]
    if not year:
        return None

    # Get first author last name from the parsed authors
    authors = citation["authors"]
    if not authors:
        return None

    first_author = authors.split(";")[0].strip()
    # Extract last name (before first comma)
    last_name = first_author.split(",")[0].strip()
    # Clean up any stray punctuation
    last_name = last_name.strip(".,;: ")

    # Try exact match on index
    key = (year, last_name.lower())
    if key in pdf_index:
        return pdf_index[key][0]

    # Try fuzzy: check if any PDF key contains the author or vice versa
    for (yr, auth), paths in pdf_index.items():
        if yr == year and (
            last_name.lower() in auth or auth in last_name.lower()
        ):
            return paths[0]

    # Last resort: scan all PDFs for year + author substring in filename
    for pdf_path in all_pdfs:
        fname = pdf_path.stem.lower()
        if year in fname and last_name.lower() in fname:
            return pdf_path

    # Try year ±1 (some citations have original vs published year mismatch)
    for offset in [-1, 1, -2, 2]:
        alt_year = str(int(year) + offset)
        alt_key = (alt_year, last_name.lower())
        if alt_key in pdf_index:
            return pdf_index[alt_key][0]
        for pdf_path in all_pdfs:
            fname = pdf_path.stem.lower()
            if alt_year in fname and last_name.lower() in fname:
                return pdf_path

    return None


def main():
    parser = argparse.ArgumentParser(description="Prepare CSV from Excel + PDF folder")
    parser.add_argument("--excel", required=True, help="Path to Excel file")
    parser.add_argument("--pdfs", required=True, help="Path to PDF folder root")
    parser.add_argument("--output", default="data/papers_metadata.csv", help="Output CSV path")
    args = parser.parse_args()

    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required. Install with: pip install openpyxl")
        sys.exit(1)

    excel_path = Path(args.excel)
    pdf_root = Path(args.pdfs)
    output_path = Path(args.output)

    print(f"Reading Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data_rows = [r for r in rows[1:] if r[0]]

    print(f"Found {len(data_rows)} citation entries")
    print(f"Columns: {header}")

    # Index PDFs
    print(f"\nScanning PDFs in: {pdf_root}")
    pdf_index, all_pdfs = find_pdfs(pdf_root)
    print(f"Found {len(all_pdfs)} PDF files")

    # Process each citation
    matched = 0
    unmatched = []
    results = []

    for row in data_rows:
        fuente = str(row[0]).strip()
        tema = str(row[1]).strip() if row[1] else ""
        subtema = str(row[2]).strip() if row[2] else ""

        citation = parse_citation(fuente)
        pdf_path = match_citation_to_pdf(citation, pdf_index, all_pdfs)

        if pdf_path:
            matched += 1
            results.append({
                "filename": str(pdf_path),
                "title": citation["title"],
                "authors": citation["authors"],
                "year": citation["year"],
                "publication": citation["publication"],
                "doi": citation["doi"],
                "abstract": "",
                "tema": tema,
                "subtema": subtema,
            })
        else:
            unmatched.append((citation.get("year", "?"), citation.get("authors", "?")[:40], fuente[:80]))

    # Check for PDFs not in the Excel
    matched_paths = {r["filename"] for r in results}
    orphan_pdfs = [p for p in all_pdfs if str(p) not in matched_paths]

    # Include orphan PDFs with metadata extracted from filename + folder structure
    orphan_included = 0
    for pdf_path in orphan_pdfs:
        fname = pdf_path.stem
        # Skip PREDATORY-tagged papers
        if "PREDATORY" in fname.upper():
            continue
        m = re.match(r"(\d{4})\w?\s+(.+)", fname)
        if m:
            yr = m.group(1)
            author_part = m.group(2)
            # Derive topic from parent folder names
            rel = pdf_path.relative_to(pdf_root)
            folder_parts = [p for p in rel.parts[:-1] if p]
            tema = folder_parts[0] if folder_parts else ""
            subtema = folder_parts[1] if len(folder_parts) > 1 else ""
            results.append({
                "filename": str(pdf_path),
                "title": f"{author_part} ({yr})",  # Placeholder title from filename
                "authors": author_part.replace(" & ", "; ").replace(" et al", ""),
                "year": yr,
                "publication": "",
                "doi": "",
                "abstract": "",
                "tema": tema,
                "subtema": subtema,
            })
            orphan_included += 1

    # Write CSV
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["filename", "title", "authors", "year", "publication", "doi", "abstract", "tema", "subtema"])
        writer.writeheader()
        writer.writerows(results)

    # Report
    print(f"\n{'='*60}")
    print(f"MATCHING SUMMARY")
    print(f"{'='*60}")
    print(f"Citations in Excel: {len(data_rows)}")
    print(f"PDFs found:         {len(all_pdfs)}")
    print(f"Matched:            {matched}")
    print(f"Unmatched citations:{len(unmatched)}")
    print(f"Orphan PDFs:        {len(orphan_pdfs)} ({orphan_included} included with filename metadata)")
    print(f"Total in CSV:       {len(results)}")
    print(f"Output CSV:         {output_path}")
    print(f"{'='*60}")

    if unmatched:
        print(f"\nUnmatched citations ({len(unmatched)}):")
        for yr, auth, cite in unmatched[:15]:
            print(f"  [{yr}] {auth} — {cite}")
        if len(unmatched) > 15:
            print(f"  ... and {len(unmatched) - 15} more")

    if orphan_pdfs:
        print(f"\nOrphan PDFs not in Excel ({len(orphan_pdfs)}):")
        for p in orphan_pdfs[:10]:
            print(f"  {p.name}")
        if len(orphan_pdfs) > 10:
            print(f"  ... and {len(orphan_pdfs) - 10} more")


if __name__ == "__main__":
    main()
